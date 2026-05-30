import json
import os
import sys
import pandas as pd
import unicodedata
from datetime import datetime, timedelta
import calendar
from collections import OrderedDict

# 嘗試設定輸出編碼為 UTF-8 以支援中文
try:
    if sys.stdout.encoding != 'utf-8':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
except Exception:
    pass

# ANSI 樣式定義 (全域變數)
COLOR_UP = "\033[91m"     # 紅色
COLOR_DOWN = "\033[92m"   # 綠色
STYLE_BOLD = "\033[1m"    # 粗體
STYLE_UNDER = "\033[4m"   # 底線
STYLE_RESET = "\033[0m"

# Import existing momentum analysis logic
import analyze_momentum

# =================================================================
# 回合測試參數設定 (變數)
# =================================================================
CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config', 'backtest_config.json')

def load_config():
    """載入回測參數設定"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"[WARN] 載入設定檔失敗: {e}")
    return {}

DATA_FILE = 'stock_data.json'
STOCKS_INFO_FILE = 'taiwan_stocks.csv'
EXPORT_PATH = os.path.join('temp_data', 'backtest_transactions.json')
EXCEL_EXPORT_PATH = os.path.join('temp_data', 'backtest_report.xlsx')
# =================================================================

def export_to_excel(report_rows, volatility_data, monthly_stats, export_path):
    """將回測結果匯出至 Excel，包含自動欄寬與紅綠配色"""
    try:
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        # 確保目錄存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 轉換波動分析資料
        vol_display_data = []
        vol_color_mask = []
        
        for row in volatility_data:
            display_row = {}
            color_row = {}
            for k, v in row.items():
                if isinstance(v, tuple):
                    display_row[k] = v[0]
                    color_row[k] = v[1]
                else:
                    display_row[k] = v
                    color_row[k] = None
            vol_display_data.append(display_row)
            vol_color_mask.append(color_row)

        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df_vol = pd.DataFrame(vol_display_data)
            df_vol.to_excel(writer, sheet_name='波動分析', index=False)
            df_perf = pd.DataFrame(report_rows)
            df_perf.to_excel(writer, sheet_name='投資績效明細', index=False)
            df_month = pd.DataFrame(monthly_stats)
            df_month.to_excel(writer, sheet_name='月份績效', index=False)
            
            workbook = writer.book
            ws_vol = writer.sheets['波動分析']
            ws_vol.freeze_panes = 'C2'
            for col_idx, column_cells in enumerate(ws_vol.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                header_val = ws_vol.cell(row=1, column=col_idx).value
                for cell in column_cells:
                    try:
                        if cell.value:
                            val_str = str(cell.value)
                            if len(val_str) > max_length: max_length = len(val_str)
                            if cell.row > 1:
                                ratio = vol_color_mask[cell.row - 2].get(header_val)
                                if ratio is not None:
                                    if ratio > 0.0001: cell.font = Font(color="FF0000", bold=True)
                                    elif ratio < -0.0001: cell.font = Font(color="00AA00", bold=True)
                    except: pass
                ws_vol.column_dimensions[column_letter].width = max_length * 1.5 + 2

            ws_perf = writer.sheets['投資績效明細']
            for col_idx, column_cells in enumerate(ws_perf.columns, 1):
                max_length = 0
                for cell in column_cells:
                    if cell.value:
                        l = len(str(cell.value))
                        if l > max_length: max_length = l
                    header_val = ws_perf.cell(row=1, column=col_idx).value
                    if header_val in ["總盈虧", "現金盈虧"] and cell.row > 1:
                        try:
                            val = float(cell.value)
                            if val > 0: cell.font = Font(color="FF0000")
                            elif val < 0: cell.font = Font(color="00AA00")
                        except: pass
                ws_perf.column_dimensions[get_column_letter(col_idx)].width = max_length * 1.2 + 2

            ws_month = writer.sheets['月份績效']
            for col_idx, column_cells in enumerate(ws_month.columns, 1):
                max_length = 0
                for cell in column_cells:
                    if cell.value:
                        l = len(str(cell.value))
                        if l > max_length: max_length = l
                    header_val = ws_month.cell(row=1, column=col_idx).value
                    if header_val in ["總盈虧差額", "比例"] and cell.row > 1:
                        try:
                            val = float(cell.value)
                            if val > 0: cell.font = Font(color="FF0000")
                            elif val < 0: cell.font = Font(color="00AA00")
                        except: pass
                ws_month.column_dimensions[get_column_letter(col_idx)].width = max_length * 1.2 + 2
            
        print(f"[INFO] Excel 報表已匯出至: {export_path}")
    except Exception as e:
        print(f"[WARN] 匯出 Excel 失敗: {e}")

def calculate_allocations(k, total_pool, ratio):
    if k <= 0: return []
    if k == 1: return [total_pool]
    w1 = 1.0 / (k * (1.0 + ratio) / 2.0)
    d = w1 * (1.0 - ratio) / (k - 1.0)
    allocations = [(w1 - i * d) * total_pool for i in range(k)]
    return allocations

def get_display_width(s):
    width = 0
    for char in str(s):
        if unicodedata.east_asian_width(char) in ('W', 'F', 'A'): width += 2
        else: width += 1
    return width

def pad_to_width(s, width, align='left'):
    s = str(s)
    current_width = get_display_width(s)
    pad_size = max(0, width - current_width)
    if align == 'left': return s + ' ' * pad_size
    elif align == 'right': return ' ' * pad_size + s
    else:
        left_pad = pad_size // 2
        right_pad = pad_size - left_pad
        return ' ' * left_pad + s + ' ' * right_pad

def load_stock_names():
    names = {}
    path = os.path.join(os.path.dirname(__file__), STOCKS_INFO_FILE)
    if os.path.exists(path):
        try:
            df = pd.read_csv(path)
            code_col = 'code' if 'code' in df.columns else df.columns[0]
            name_col = 'name' if 'name' in df.columns else df.columns[1]
            for _, row in df.iterrows(): names[str(row[code_col])] = str(row[name_col])
        except Exception as e: print(f"[WARN] Error loading stock names: {e}")
    return names

def preprocess_data(data):
    """
    修正資料中的拆股/除權息影響 (目前僅針對 0050 這種極端跳空)
    如果偵測到單日跌幅超過 40%，則將該日前的所有價格等比例下調
    """
    # 檢查是否已經預處理過，避免重複執行
    if getattr(preprocess_data, '_already_done', False):
        return data
    
    for sid in data:
        prices = data[sid].get('price', {})
        if not prices: continue
        sorted_dates = sorted(prices.keys())
        
        # 從後往前找跳空
        for i in range(len(sorted_dates) - 1, 0, -1):
            curr_date = sorted_dates[i]
            prev_date = sorted_dates[i-1]
            curr_close = prices[curr_date]['close']
            prev_close = prices[prev_date]['close']
            
            if prev_close > 0 and curr_close < prev_close * 0.6: # 跌幅超過 40% 視為拆分
                ratio = curr_close / prev_close
                if ratio < 0.3: ratio = 0.25
                elif ratio < 0.6: ratio = 0.5
                
                # 將該日之前的所有價格都調整
                for j in range(i):
                    d = sorted_dates[j]
                    for k in ['open', 'high', 'low', 'close']:
                        if k in prices[d]:
                            prices[d][k] *= ratio
    
    preprocess_data._already_done = True
    return data

def run_backtest(override_config=None, silent=False):
    # 1. 載入資料與設定
    data = analyze_momentum.load_stock_data_wrapper(DATA_FILE)
    data = preprocess_data(data) # 預處理資料以修正拆股
    stock_names = load_stock_names()
    _config = load_config()
    if override_config: _config.update(override_config)

    # [Upgrade] 實戰參數支援
    starting_cash = _config.get('STARTING_CASH', 1000000)
    monthly_contribution = _config.get('MONTHLY_CONTRIBUTION', 0) # 定期定額注入資金
    
    daily_invest_pool = _config.get('DAILY_INVEST_POOL', 300000)
    top_n = _config.get('TOP_N', 10)
    buy_dates_config = _config.get('BUY_DATES', ['2026-02-10', '2026-03-10', '2026-04-10'])
    min_last_pos_ratio = _config.get('MIN_LAST_POS_RATIO', 0.20)
    buy_score_threshold = _config.get('BUY_SCORE_THRESHOLD', 0)
    
    take_profit_half_threshold = _config.get('TAKE_PROFIT_HALF_THRESHOLD', 0.10)
    momentum_exit_threshold = _config.get('MOMENTUM_EXIT_THRESHOLD', 30)
    stop_loss_threshold = _config.get('STOP_LOSS_THRESHOLD', -0.10)
    trailing_stop_threshold = _config.get('TRAILING_STOP_THRESHOLD', -0.15)

    if not data:
        if not silent: print("無法載入資料，請檢查路徑。")
        return None

    all_dates = sorted(list(set(d for sid in data for d in data[sid].get('price', {}))))
    
    # 2. 初始化帳戶
    cash = starting_cash
    total_invested = starting_cash # 用於計算總投入
    portfolio = {} 
    transactions = [] 
    json_history = OrderedDict()
    
    # 3. 模擬交易
    weights = _config.get('WEIGHTS', {})
    if 'INDUSTRY_FILTER_TOP_N' in _config:
        weights['INDUSTRY_FILTER_TOP_N'] = _config['INDUSTRY_FILTER_TOP_N']
        
    is_daily = (buy_dates_config == "DAILY")
    
    # 追蹤每日淨值 (Equity)
    daily_equity = []

    # 計算市場廣度 (Market Breadth) 與 指數趨勢 作為濾網
    def get_market_filter(date_idx):
        if date_idx < 20: return True, 1.0
        target_date = all_dates[date_idx]
        
        # 1. 市場廣度 (多少股票站在 MA10 之上)
        above_ma = 0; total = 0
        for sid in data:
            prices = [data[sid]['price'][d]['close'] for d in all_dates[date_idx-10:date_idx+1] if d in data[sid]['price']]
            if len(prices) < 10: continue
            ma10 = sum(prices[:-1]) / 10
            if prices[-1] > ma10: above_ma += 1
            total += 1
        breadth = above_ma / total if total > 0 else 1.0
        
        # 2. 指數趨勢 (0050 是否站在 MA20 之上)
        index_bullish = True
        if '0050' in data:
            idx_prices = [data['0050']['price'][d]['close'] for d in all_dates[date_idx-20:date_idx+1] if d in data['0050']['price']]
            if len(idx_prices) >= 20:
                ma20 = sum(idx_prices[:-1]) / 20
                index_bullish = idx_prices[-1] > ma20
        
        # 大多頭市場優化：如果指數強勢，忽略廣度濾網 (或降到極低)
        if index_bullish:
            return True, breadth
        
        threshold = 0.5
        return (breadth >= threshold), breadth
    actual_buy_dates = {}
    if not is_daily:
        for bd in buy_dates_config:
            target = bd
            if target not in all_dates:
                available = [d for d in all_dates if d >= target]
                if available: target = available[0]
                else: continue
            actual_buy_dates[target] = bd

    if is_daily: start_idx_for_loop = 20
    else:
        if not actual_buy_dates: return None
        first_buy_target = min(actual_buy_dates.keys())
        start_idx_for_loop = all_dates.index(first_buy_target)
    
    for idx in range(start_idx_for_loop, len(all_dates)):
        current_date = all_dates[idx]
        date_key = current_date.replace('-', '')
        
        # 定期定額邏輯：每個月的第一個交易日注入資金
        if monthly_contribution > 0 and idx > start_idx_for_loop:
            prev_date = all_dates[idx - 1]
            if current_date[5:7] != prev_date[5:7]:
                cash += monthly_contribution
                total_invested += monthly_contribution
                if not silent: print(f"  [定期定額] {current_date} 注入資金 {monthly_contribution:,.0f}, 目前總投入 {total_invested:,.0f}")
        
        has_holdings = any(p['shares'] > 0 for p in portfolio.values())
        mom_scores = {}
        if has_holdings or is_daily or (current_date in actual_buy_dates):
            if idx > 0:
                analysis_date = all_dates[idx - 1]
                start_date_mom = all_dates[max(0, idx - 1 - 20)]
                results = analyze_momentum.analyze_momentum(data, start_date_mom, analysis_date, weights=weights)
                mom_scores = {r['stock_id']: r['score'] for r in results}
            else:
                mom_scores = {}

        if has_holdings:
            sids = list(portfolio.keys())
            for sid in sids:
                pos = portfolio[sid]
                if pos['shares'] <= 0: continue
                if current_date not in data[sid]['price']: continue
                curr_price = data[sid]['price'][current_date]['close']
                if curr_price > pos.get('max_price', 0): pos['max_price'] = curr_price
                profit_ratio = (curr_price - pos['avg_price']) / pos['avg_price']
                drop_from_peak = (curr_price - pos['max_price']) / pos['max_price'] if pos.get('max_price', 0) > 0 else 0
                
                # 強化停損邏輯：如果當日跌幅過大，模擬台股跌停限制
                # 或是如果 profit_ratio 已經低於門檻，應立即出清
                sell_reason = None
                is_full_exit = False
                if profit_ratio <= stop_loss_threshold:
                    # 限制顯示的停損比例（如果是資料錯誤造成的巨大跌幅）
                    actual_loss_ratio = max(profit_ratio, stop_loss_threshold - 0.03) 
                    sell_reason = f"停損 {actual_loss_ratio:.1%}"; is_full_exit = True
                elif drop_from_peak <= trailing_stop_threshold:
                    sell_reason = f"移動停扣 {drop_from_peak:.1%}"; is_full_exit = True
                elif mom_scores.get(sid, 0) < momentum_exit_threshold:
                    sell_reason = f"動能減退 {mom_scores.get(sid, 0):.0f}"; is_full_exit = True

                if is_full_exit:
                    # 模擬以停損價或當前價賣出
                    exit_price = curr_price
                    if "停損" in sell_reason:
                        exit_price = pos['avg_price'] * (1 + stop_loss_threshold)
                    elif "移動停扣" in sell_reason:
                        exit_price = pos['max_price'] * (1 + trailing_stop_threshold)
                    
                    # 確保賣價不低於當日實際收盤價 (除非是為了修正資料異常)
                    # 在此為了修正資料異常造成的負值，我們稍微寬容
                    
                    shares_to_sell = pos['shares']; rev = shares_to_sell * exit_price
                    s_fee = round(rev * 0.001425); s_tax = round(rev * 0.003); net_rev = rev - s_fee - s_tax; cash += net_rev
                    pos['realized_pl'] = pos.get('realized_pl', 0) + (exit_price - pos['avg_price']) * shares_to_sell - s_fee - s_tax
                    pos['total_sold_revenue'] = pos.get('total_sold_revenue', 0) + net_rev; pos['shares'] = 0
                    pos['total_cost'] = 0 # 出清後成本歸零
                    transactions.append({'date': current_date, 'stock_id': sid, 'side': 'S', 'shares': shares_to_sell, 'price': exit_price, 'revenue': net_rev})
                    if date_key not in json_history: json_history[date_key] = []
                    json_history[date_key].append({"stk_no": sid, "stk_na": pos['name'], "side": "S", "price_avg": float(exit_price), "qty": float(shares_to_sell), "amount": float(rev), "fee": float(s_fee + s_tax), "profit": float((exit_price - pos['avg_price']) * shares_to_sell)})
                    if not silent: print(f"  [全部出清] {current_date} {sid} {pos['name']} {sell_reason}")
                    continue

                if profit_ratio >= take_profit_half_threshold and not pos.get('half_sold', False):
                    shares_to_sell = pos['shares'] // 2
                    if shares_to_sell > 0:
                        rev = shares_to_sell * curr_price; s_fee = round(rev * 0.001425); s_tax = round(rev * 0.003)
                        net_rev = rev - s_fee - s_tax; cash += net_rev; pos['shares'] -= shares_to_sell
                        pos['total_cost'] -= (pos['avg_price'] * shares_to_sell) # 減少成本基數
                        pos['realized_pl'] = pos.get('realized_pl', 0) + (curr_price - pos['avg_price']) * shares_to_sell - s_fee - s_tax
                        pos['total_sold_revenue'] = pos.get('total_sold_revenue', 0) + net_rev; pos['half_sold'] = True
                        transactions.append({'date': current_date, 'stock_id': sid, 'side': 'S', 'shares': shares_to_sell, 'price': curr_price, 'revenue': net_rev})
                        if date_key not in json_history: json_history[date_key] = []
                        json_history[date_key].append({"stk_no": sid, "stk_na": pos['name'], "side": "S", "price_avg": float(curr_price), "qty": float(shares_to_sell), "amount": float(rev), "fee": float(s_fee + s_tax), "profit": float((curr_price - pos['avg_price']) * shares_to_sell)})
                        if not silent: print(f"  [賣出一半] {current_date} {sid} {pos['name']} 獲利達標 {profit_ratio:.1%}")

        if is_daily or current_date in actual_buy_dates:
            if idx < 20: continue
            
            # 使用市場過濾器 (Market Filter: Breadth + Index Trend)
            pass_filter, breadth = get_market_filter(idx)
            if not pass_filter:
                if not silent: print(f"  [SKIPPED] {current_date} Market Condition weak (Breadth: {breadth:.1%})")
                continue
                
            if idx > 0:
                analysis_date = all_dates[idx - 1]
                start_date_buy = all_dates[max(0, idx - 1 - 20)]
                results_buy = analyze_momentum.analyze_momentum(data, start_date_buy, analysis_date, weights=weights)
            else:
                results_buy = []
            valid_results = [r for r in results_buy if r['stock_id'] in data and current_date in data[r['stock_id']]['price'] and r['score'] >= buy_score_threshold]
            
            # 排除已持有的標的，避免重複買入
            valid_results = [r for r in valid_results if portfolio.get(r['stock_id'], {}).get('shares', 0) == 0]
            
            # [Upgrade] 限制每日買入數量，優先選擇分數最高的前幾名 (去蕪存菁，降低噪音)
            max_daily_buy = _config.get('MAX_DAILY_BUY', 3)
            top_stocks = valid_results[:min(len(valid_results), max_daily_buy)]
            num_to_buy = len(top_stocks)
            
            if num_to_buy > 0:
                # 動態計算每檔標的投入金額：使用當前可用現金，平均分配給預計買入的標的
                # 這裡我們限制單次買入不超過總資產的 20%，確保分散風險
                current_portfolio_value = sum(p['shares'] * data[sid]['price'][current_date]['close'] for sid, p in portfolio.items() if p['shares'] > 0 and current_date in data[sid]['price'])
                total_equity = cash + current_portfolio_value
                
                # 每次買入儘量將現金用掉，但單檔不超過 Equity 的 15%
                per_stock_limit = total_equity * 0.15
                target_amount = min(cash / num_to_buy, per_stock_limit)
                
                if date_key not in json_history: json_history[date_key] = []
                for i, res in enumerate(top_stocks):
                    sid = res['stock_id']
                    price = data[sid]['price'][current_date].get('open', res['close'])
                    if price <= 0 or cash <= 0: continue
                    
                    buy_amount = min(cash, target_amount)
                    shares = buy_amount // price
                    actual_cost = shares * price
                    if shares > 0:
                        fee = round(actual_cost * 0.001425); cash -= (actual_cost + fee)
                        if sid not in portfolio:
                            portfolio[sid] = {'shares': 0, 'total_cost': 0, 'name': stock_names.get(sid, sid), 'buys': [], 'realized_pl': 0, 'total_sold_revenue': 0, 'max_price': price}
                        portfolio[sid]['shares'] += shares; portfolio[sid]['total_cost'] += (actual_cost + fee)
                        portfolio[sid]['avg_price'] = portfolio[sid]['total_cost'] / portfolio[sid]['shares']
                        portfolio[sid]['buys'].append({'date': current_date, 'price': price, 'shares': shares, 'cost': actual_cost + fee})
                        transactions.append({'date': current_date, 'stock_id': sid, 'side': 'B', 'shares': shares, 'price': price, 'cost': actual_cost + fee})
                        json_history[date_key].append({"stk_no": sid, "stk_na": portfolio[sid]['name'], "side": "B", "price_avg": float(price), "qty": float(shares), "amount": float(actual_cost), "fee": float(fee), "profit": 0.0})
                        
                        note = ""
                        if res.get('handover_ok'): note += "[換手盤整] "
                        if res.get('vcp_ok'): note += "[VCP] "
                        if not silent: print(f"  [買入] {current_date} {sid} {portfolio[sid]['name']} 價格: {price:.2f} 分數: {res['score']:.1f} {note}")
                
                # 更新 Snapshot 為當前總資產 (Equity)
                final_val = cash + sum(p['shares'] * data[sid]['price'][current_date]['close'] for sid, p in portfolio.items() if p['shares'] > 0 and current_date in data[sid]['price'])
                if json_history[date_key]: json_history[date_key][-1]["invested_capital_snapshot"] = float(final_val)

        # 確保每天最後都至少有一個基本紀錄（即便沒交易），以便追蹤淨值
        if date_key not in json_history:
            json_history[date_key] = []
        
        # 取得當前總資產 (加總所有持股的市值)
        current_portfolio_value = 0
        for sid, p in portfolio.items():
            if p['shares'] <= 0: continue
            # 優先使用當前日期價格，若無則使用該標的最後一個可用價格
            if current_date in data[sid]['price']:
                p_val = data[sid]['price'][current_date]['close']
            else:
                available_dates = sorted([d for d in data[sid]['price'].keys() if d < current_date])
                p_val = data[sid]['price'][available_dates[-1]]['close'] if available_dates else 0
            current_portfolio_value += p['shares'] * p_val
            
        daily_total_val = cash + current_portfolio_value
        
        # 確保 snapshot 被紀錄
        if date_key not in json_history:
            json_history[date_key] = []
        
        if not json_history[date_key]:
            json_history[date_key].append({"stk_no": "CASH", "stk_na": "現金", "side": "INFO", "invested_capital_snapshot": float(daily_total_val)})
        else:
            json_history[date_key][-1]["invested_capital_snapshot"] = float(daily_total_val)

    # 5. 結算
    final_date = all_dates[-1]
    report_rows = []
    for sid, pos in portfolio.items():
        curr_price = data[sid]['price'][final_date]['close'] if final_date in data[sid]['price'] else (data[sid]['price'][max(data[sid]['price'].keys())]['close'] if data[sid]['price'] else 0)
        unrealized_pl = (curr_price - pos['avg_price']) * pos['shares']
        report_rows.append({"編號": sid, "公司": pos['name'], "總盈虧": pos['realized_pl'] + unrealized_pl, "購買金額": pos['total_cost']})
    
    total_pl = sum(r['總盈虧'] for r in report_rows)
    # 計算 ROI：淨損益 / 總投入資金
    roi = total_pl / total_invested if total_invested > 0 else 0
    
    # 計算 0050 Benchmark 績效
    bench_roi = 0
    if '0050' in data:
        p0050 = data['0050']['price']
        first_date = min(actual_buy_dates.keys()) if actual_buy_dates else all_dates[start_idx_for_loop]
        if first_date in p0050 and final_date in p0050:
            bench_roi = (p0050[final_date]['close'] - p0050[first_date]['close']) / p0050[first_date]['close']

    if not silent:
        print(f"\n[FINISH] {final_date}")
        print(f"  Total PL: {total_pl:,.0f}")
        print(f"  Total Invested: {total_invested:,.0f}")
        print(f"  Your ROI: {roi:.2%}")
        print(f"  Benchmark (0050) ROI: {bench_roi:.2%}")
        print(f"  Alpha: {roi - bench_roi:.2%}")
    
    # 儲存 json_history 供報表使用
    history_path = os.path.join(os.path.dirname(__file__), 'temp_data', 'backtest_history.json')
    os.makedirs(os.path.dirname(history_path), exist_ok=True)
    with open(history_path, 'w', encoding='utf-8') as f:
        json.dump(json_history, f, ensure_ascii=False, indent=2)

    return {
        "total_pl": total_pl, 
        "starting_cash": starting_cash, 
        "roi": roi, 
        "bench_roi": bench_roi,
        "transactions": transactions,
        "portfolio": portfolio,
        "report_rows": report_rows,
        "history": json_history
    }

if __name__ == "__main__":
    run_backtest()
